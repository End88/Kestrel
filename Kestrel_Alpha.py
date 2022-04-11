import openpyxl
import shutil
import tempfile

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    #print(wb)

    h = 'none'  # Nome do arquivo
    funct_1 = 0 # Salva variável 1
    funct_2 = 0 # Salva variável 2
    number = 0  # Salva número de controle
    name = 0

    with open('Instruções_Kestrel.txt', 'r') as arquivo, \
            tempfile.NamedTemporaryFile('w', delete=False) as out:
        for index, linha in enumerate(arquivo, start=1):
            if index == 6:
                h = linha
            if index == 10:
                funct_1 = int(linha)

            if index == 15:  # linha 6, mudar o conteúdo
                number = int(linha)
                array = list(linha)
                if int(array[-2]) == funct_1:
                    number += 1
                else:
                    number += 9

                out.write(str(number) + "\n")
            else:  # não é linha 6, escreve a linha sem modificação
                out.write(linha)

    # move o arquivo temporário para o original
    shutil.move(out.name, 'Instruções_Kestrel.txt')

    array = list(h)
    del(array[-1]) # Ajuste de nome de arquivo, exclusão de \n
    h = "".join(array)

    wb = openpyxl.load_workbook(h)

    ws = wb['xxxx']
    ws.title = str(number)
    img = openpyxl.drawing.image.Image('img\\logo1.png')
    img.anchor = 'A1'
    ws['D10'] = number
    ws.add_image(img)

    ws = wb['DADOS DO PROCESSO']
    img = openpyxl.drawing.image.Image('img\\logo2.png')
    img.anchor = 'A1'
    ws.add_image(img)
    ws['B7'] = number
    ws['A27'] = number
    wb.save(str(number) + ' - ' + '.xlsx')

